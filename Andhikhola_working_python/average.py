# -*- coding: utf-8 -*-
"""
Created on Wed Jan 18 09:41:18 2023

@author: kumar
"""

import openpyxl

# Open the existing Excel file
workbook = openpyxl.load_workbook('Andhikhola_foraverage.xlsx')
sheet = workbook['61-62']

# Create a new sheet for the averages
averages_sheet = workbook.create_sheet('Averages')

# Calculate the averages for each column in range B7:M38
for column in range(2, 14):
    column_range = f"B{column}:{chr(ord('B') + column - 2)}38"
    column_averages = [cell.value for cell in sheet[column_range] if cell.value is not None]
    average = sum(column_averages) / len(column_averages)
    averages_sheet.cell(row=1, column=column).value = average

# Save the new sheet to the same Excel file
workbook.save('Andhikhola_foraverage.xlsx')


